import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date
import json
import os

DATA_FILE = "patients_data.json"

ICD_CODES = {
    "1": ("Z00.00", "General Health Checkup"),
    "2": ("J06.9",  "Upper Respiratory Infection"),
    "3": ("E11.9",  "Type 2 Diabetes"),
    "4": ("I10",    "Hypertension"),
    "5": ("M54.5",  "Lower Back Pain"),
    "6": ("J18.9",  "Pneumonia"),
    "7": ("K21.0",  "Acid Reflux / GERD"),
    "8": ("N39.0",  "Urinary Tract Infection"),
}

CPT_CODES = {
    "1": ("99201", "Office Visit - New Patient",         500),
    "2": ("99211", "Office Visit - Established Patient", 300),
    "3": ("93000", "ECG / EKG",                          800),
    "4": ("85025", "Complete Blood Count (CBC)",         400),
    "5": ("71046", "Chest X-Ray",                       1200),
    "6": ("99283", "Emergency Visit - Moderate",        2500),
}

# ── LOAD & SAVE ──
def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    return []

def save_data(patients):
    with open(DATA_FILE, "w") as f:
        json.dump(patients, f, indent=4)

def show_menu():
    print("\n" + "="*50)
    print("   MEDICAL BILLING MANAGEMENT SYSTEM")
    print("="*50)
    print("  1. Add New Patient & Bill")
    print("  2. View All Records")
    print("  3. Update Claim Status")
    print("  4. Export to Excel")
    print("  5. Exit")
    print("="*50)

def select_icd():
    print("\n── SELECT DIAGNOSIS (ICD-10) ──")
    for key, (code, desc) in ICD_CODES.items():
        print(f"  {key}. [{code}]  {desc}")
    choice = input("Enter number: ").strip()
    if choice in ICD_CODES:
        return ICD_CODES[choice]
    return ("Z00.00", "General Health Checkup")

def select_cpt():
    print("\n── SELECT PROCEDURE (CPT) ──")
    for key, (code, desc, fee) in CPT_CODES.items():
        print(f"  {key}. [{code}]  {desc}  —  Rs. {fee}")
    choice = input("Enter number: ").strip()
    if choice in CPT_CODES:
        return CPT_CODES[choice]
    return ("99201", "Office Visit - New Patient", 500)

def add_patient(patients):
    print("\n── NEW PATIENT ENTRY ──")
    name = input("Patient Name       : ").strip()
    if not name:
        print("Error: Name cannot be empty.")
        return
    age       = input("Age                : ").strip()
    gender    = input("Gender (M/F)       : ").strip().upper()
    insurance = input("Insurance Provider : ").strip()

    icd_code, icd_desc      = select_icd()
    cpt_code, cpt_desc, fee = select_cpt()

    record = {
        "Patient Name"  : name,
        "Age"           : age,
        "Gender"        : gender,
        "Insurance"     : insurance,
        "ICD-10 Code"   : icd_code,
        "Diagnosis"     : icd_desc,
        "CPT Code"      : cpt_code,
        "Procedure"     : cpt_desc,
        "Amount (Rs.)"  : fee,
        "Claim Status"  : "Pending",
        "Date"          : str(date.today()),
    }

    patients.append(record)
    save_data(patients)
    print(f"\n  ✔ Saved — {name} | Bill: Rs. {fee} | Status: Pending")

def view_records(patients):
    if not patients:
        print("\n  No records found.")
        return
    print(f"\n  {'#':<4} {'Name':<20} {'ICD':<8} {'CPT':<8} {'Amount':>8}  {'Status':<12} {'Date'}")
    print("  " + "─"*75)
    for i, p in enumerate(patients, 1):
        print(f"  {i:<4} {p['Patient Name']:<20} {p['ICD-10 Code']:<8} "
              f"{p['CPT Code']:<8} Rs.{p['Amount (Rs.)']:>5}  "
              f"{p['Claim Status']:<12} {p['Date']}")
    print("  " + "─"*75)
    total = sum(p['Amount (Rs.)'] for p in patients)
    print(f"  {'TOTAL':>44}  Rs.{total:>5}")

def update_status(patients):
    if not patients:
        print("\n  No records found.")
        return
    view_records(patients)
    print("\n── UPDATE CLAIM STATUS ──")
    try:
        num = int(input("  Enter patient number: ").strip())
        if 1 <= num <= len(patients):
            patient = patients[num - 1]
            print(f"\n  Patient : {patient['Patient Name']}")
            print(f"  Current : {patient['Claim Status']}")
            print("\n  1. Paid")
            print("  2. Pending")
            print("  3. Rejected")
            choice = input("  New status: ").strip()
            status_map = {"1": "Paid", "2": "Pending", "3": "Rejected"}
            if choice in status_map:
                old = patient['Claim Status']
                patient['Claim Status'] = status_map[choice]
                save_data(patients)
                print(f"\n  ✔ {patient['Patient Name']}: {old} → {patient['Claim Status']}")
            else:
                print("  Invalid choice.")
        else:
            print("  Invalid number.")
    except ValueError:
        print("  Enter a valid number.")

def export_to_excel(patients):
    if not patients:
        print("\n  No records to export.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Billing Records"

    headers = list(patients[0].keys())
    header_fill = PatternFill("solid", fgColor="1B3A6B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_i, record in enumerate(patients, 2):
        fill_color = "EFF4FD" if row_i % 2 == 0 else "FFFFFF"
        row_fill = PatternFill("solid", fgColor=fill_color)
        for col_i, value in enumerate(record.values(), 1):
            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center")
            if value == "Paid":
                cell.font = Font(bold=True, color="166534")
            elif value == "Rejected":
                cell.font = Font(bold=True, color="991B1B")
            elif value == "Pending":
                cell.font = Font(bold=True, color="92400E")

    total_row = len(patients) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    total = sum(p['Amount (Rs.)'] for p in patients)
    ws.cell(row=total_row, column=9, value=total).font = Font(bold=True, color="1B3A6B")

    widths = [20, 5, 8, 18, 12, 26, 10, 28, 14, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    filename = f"billing_report_{date.today()}.xlsx"
    wb.save(filename)
    print(f"\n  ✔ Excel saved: {filename}")

def main():
    patients = load_data()
    print(f"\n  Welcome — {len(patients)} existing record(s) loaded.")
    while True:
        show_menu()
        choice = input("  Select option: ").strip()
        if   choice == "1": add_patient(patients)
        elif choice == "2": view_records(patients)
        elif choice == "3": update_status(patients)
        elif choice == "4": export_to_excel(patients)
        elif choice == "5":
            print("\n  Goodbye!\n")
            break
        else:
            print("  Invalid option.")

main()